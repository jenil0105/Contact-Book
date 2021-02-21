import os
import openpyxl
from prettytable import PrettyTable

def choice():
    def load_workbook(wb_path):
        if os.path.exists(wb_path):
            return openpyxl.load_workbook(wb_path)
        return openpyxl.Workbook()

    wb_path = 'contact book.xlsx'
    wb = load_workbook(wb_path)
    sheet = wb['Sheet']
    
    heading=['Name','Mobile Number','Email-id']

    for col_idx, title in enumerate(heading):
        sheet.cell(row=1, column=col_idx+1, value=title)
    wb.save(wb_path)
    
    class Contact():
        
        def Add(self):
            
            data=[eval(i) for i in input('Enter name, mobile no., mailid:').split(',')]
            for row in range(1,sheet.max_row+1):
                if sheet.cell(row,column=1).value==data[0]: 
                               print('\nName is already available\n')
                               break
            else:
                sheet.append(data)
                wb.save(wb_path)
                print('\nSuccessfully data is added')

        def Display(self):
            
            myTable=PrettyTable(['Name','Mobile Number','Email-id'])
            list_display=[]
            
            for row in range(2,sheet.max_row+1):
                for column in range(1,sheet.max_column+1):
                    d= sheet.cell(row,column).value
                    list_display.append(d)
                myTable.add_row(list_display)
                list_display=[]
            print(myTable)
        
        def Search(self):
            
            myTable=PrettyTable(['Name','Mobile Number','Email-id'])
            list_search=[]
            
            search=input('\nEnter the name you want to search:')
            
            for row in range(1,sheet.max_row+1):
                if sheet.cell(row,column=1).value==search:  
                    for column in range(1,sheet.max_column+1):
                        s= sheet.cell(row,column).value
                        list_search.append(s)
                    myTable.add_row(list_search)
                    list_search=[]
                    print('\nData is\n',myTable)
                    break
            else:
                print('\nData not found')
                
        def Delete(self):
            
            delete=input('\nEnter the name you want to delete:')
            
            for row in range(1,sheet.max_row+1):
                if sheet.cell(row,column=1).value==delete:
                    sheet.delete_rows(row)
                    wb.save(wb_path)
                    print('\nSuccessfully deleted data')
                    break
            else:
                print('\nData not found\n')

    data=Contact()
    value = True
    while value:
        option=input('\nSelect the following options: \n 1. Add contact\n 2. Display contact\n 3. Search contact\n 4. Delete contact\n 5. Exit\n')

        if option=='1':
            data.Add()
            
        elif option=='2':
            data.Display()

        elif option=='3':
            data.Search() 

        elif option=='4':
            data.Delete()

        elif option=='5':
            break
            
        else:
            print('Choose the correct option\n')
choice()
